"""Document management API for Knowledge Base system.

Provides file upload, text extraction, chunking, embedding, and
semantic search over document collections using Qdrant vector store.
"""

import io
import logging
import os
import uuid
from datetime import UTC, datetime
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel

from app.core.embedding import get_embedding_provider
from app.core.input_sanitizer import validate_file_upload
from app.core.security import get_current_user
from app.core.vector import (
    delete_collection,
    delete_point,
    list_collections,
    recreate_collection,
    upsert_point,
    vector_store_health,
)
from app.core.vector import (
    scroll as vector_scroll,
)
from app.core.vector import (
    search as vector_search,
)
from app.models import User

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/documents", tags=["documents"])

KB_COLLECTION = "knowledge_base"
DOCUMENTS_INDEX = "documents"

# ── Schemas ──────────────────────────────────────────────────────────


class DocumentResponse(BaseModel):
    id: str
    filename: str
    title: str
    file_type: str
    size_bytes: int
    chunk_count: int
    created_at: str
    status: str


class ChunkResponse(BaseModel):
    id: str
    document_id: str
    content: str
    chunk_index: int
    score: float | None = None


class SearchRequest(BaseModel):
    query: str
    collection: str = KB_COLLECTION
    limit: int = 10
    score_threshold: float | None = 0.5
    filters: dict[str, Any] | None = None


class SearchResult(BaseModel):
    chunk: ChunkResponse
    document: dict[str, Any] | None = None


# ── Text Extraction ─────────────────────────────────────────────────


def extract_text(filename: str, content: bytes) -> str:
    ext = os.path.splitext(filename)[1].lower()
    try:
        if ext == ".pdf":
            return _extract_pdf(content)
        elif ext == ".docx":
            return _extract_docx(content)
        elif ext == ".xlsx":
            return _extract_xlsx(content)
        elif ext in (".txt", ".md", ".csv", ".json"):
            return content.decode("utf-8", errors="replace")
        else:
            return content.decode("utf-8", errors="replace")
    except Exception as e:
        logger.error("Text extraction failed for %s: %s", filename, e)
        raise HTTPException(status_code=400, detail=f"Failed to extract text: {e}") from e


def _extract_pdf(content: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    return "\n\n".join(page.extract_text() or "" for page in reader.pages)


def _extract_docx(content: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(content))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_xlsx(content: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True)
    texts = []
    for sheet in wb.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            vals = [str(c) for c in row if c is not None]
            if vals:
                rows.append(" | ".join(vals))
        if rows:
            texts.append(f"--- Sheet: {sheet.title} ---\n" + "\n".join(rows))
    return "\n\n".join(texts)


def chunk_text(text: str, chunk_size: int = 512, overlap: int = 64) -> list[str]:
    from langchain_text_splitters import RecursiveCharacterTextSplitter

    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=overlap,
        separators=["\n\n", "\n", ". ", " ", ""],
    )
    return splitter.split_text(text)


async def _ensure_kb_collection_async():
    from app.core.embedding import get_embedding_provider
    from app.core.vector import get_client

    client = get_client()
    collections = client.get_collections().collections
    names = [c.name for c in collections]
    if KB_COLLECTION not in names:
        emb = get_embedding_provider()
        from qdrant_client.http import models as qmodels

        client.create_collection(
            collection_name=KB_COLLECTION,
            vectors_config=qmodels.VectorParams(
                size=emb.dimension,
                distance=qmodels.Distance.COSINE,
            ),
        )


def _ensure_kb_collection():
    try:
        import anyio

        anyio.run(_ensure_kb_collection_async)
    except Exception:
        pass


# ── Endpoints ─────────────────────────────────────────────────────────


@router.get("/health")
async def health():
    return await vector_store_health()


@router.get("/collections")
async def get_collections(current_user: User = Depends(get_current_user)):
    cols = await list_collections()
    return {"collections": cols}


@router.post("/collections/create")
async def create_collection_endpoint(name: str = KB_COLLECTION, current_user: User = Depends(get_current_user)):
    await recreate_collection(name)
    return {"success": True, "collection": name}


@router.delete("/collections/{name}")
async def remove_collection(name: str, current_user: User = Depends(get_current_user)):
    await delete_collection(name)
    return {"success": True, "collection": name}


@router.post("/upload")
async def upload_document(
    file: UploadFile = File(...),
    title: str = Form(""),
    collection: str = Form(KB_COLLECTION),
    chunk_size: int = Form(512),
    chunk_overlap: int = Form(64),
    current_user: User = Depends(get_current_user),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    content = await file.read()
    validation = validate_file_upload(file.filename, content)
    if not validation["valid"]:
        raise HTTPException(status_code=400, detail=validation["reason"])

    doc_id = str(uuid.uuid4())
    text = extract_text(file.filename, content)
    chunks = chunk_text(text, chunk_size=chunk_size, overlap=chunk_overlap)

    emb = get_embedding_provider()
    collection_name = collection

    # Ensure collection exists
    from qdrant_client.http import models as qmodels

    from app.core.vector import get_client

    client = get_client()
    try:
        client.get_collection(collection_name=collection_name)
    except Exception:
        client.create_collection(
            collection_name=collection_name,
            vectors_config=qmodels.VectorParams(
                size=emb.dimension,
                distance=qmodels.Distance.COSINE,
            ),
        )

    # Store each chunk as a point
    doc_meta = {
        "document_id": doc_id,
        "filename": file.filename,
        "title": title or file.filename,
        "file_type": os.path.splitext(file.filename)[1].lower(),
        "size_bytes": len(content),
        "created_at": datetime.now(UTC).isoformat(),
        "user_id": str(current_user.id),
        "type": "chunk",
    }

    for i, chunk_content in enumerate(chunks):
        chunk_id = f"{doc_id}_chunk_{i}"
        payload = {**doc_meta, "chunk_index": i, "content": chunk_content}
        await upsert_point(
            collection=collection_name,
            point_id=chunk_id,
            text=chunk_content,
            payload=payload,
        )

    # Store document metadata point
    doc_payload = {
        "document_id": doc_id,
        "filename": file.filename,
        "title": title or file.filename,
        "file_type": os.path.splitext(file.filename)[1].lower(),
        "size_bytes": len(content),
        "chunk_count": len(chunks),
        "created_at": datetime.now(UTC).isoformat(),
        "user_id": str(current_user.id),
        "status": "indexed",
        "type": "document_meta",
    }
    await upsert_point(
        collection=DOCUMENTS_INDEX,
        point_id=doc_id,
        text=f"{title or file.filename} {file.filename}",
        payload=doc_payload,
    )

    return {
        "success": True,
        "document_id": doc_id,
        "filename": file.filename,
        "chunk_count": len(chunks),
        "status": "indexed",
    }


@router.get("")
async def list_documents(
    collection: str = DOCUMENTS_INDEX,
    limit: int = 100,
    offset: str | None = None,
    current_user: User = Depends(get_current_user),
):
    points, next_offset = await vector_scroll(
        collection=collection,
        filters={"user_id": str(current_user.id), "type": "document_meta"},
        limit=limit,
        offset=offset,
    )
    docs = []
    for p in points:
        pl = p["payload"]
        docs.append(
            DocumentResponse(
                id=pl.get("document_id", p["id"]),
                filename=pl.get("filename", ""),
                title=pl.get("title", ""),
                file_type=pl.get("file_type", ""),
                size_bytes=pl.get("size_bytes", 0),
                chunk_count=pl.get("chunk_count", 0),
                created_at=pl.get("created_at", ""),
                status=pl.get("status", "unknown"),
            )
        )
    return {"success": True, "documents": docs, "next_offset": next_offset, "total": len(docs)}


@router.get("/{document_id}")
async def get_document(
    document_id: str,
    collection: str = DOCUMENTS_INDEX,
    current_user: User = Depends(get_current_user),
):
    from app.core.vector import get_client

    client = get_client()
    points = client.retrieve(
        collection_name=collection,
        ids=[document_id],
        with_payload=True,
        with_vectors=False,
    )
    if not points:
        raise HTTPException(status_code=404, detail="Document not found")
    pl = points[0].payload or {}
    return {
        "success": True,
        "document": DocumentResponse(
            id=pl.get("document_id", document_id),
            filename=pl.get("filename", ""),
            title=pl.get("title", ""),
            file_type=pl.get("file_type", ""),
            size_bytes=pl.get("size_bytes", 0),
            chunk_count=pl.get("chunk_count", 0),
            created_at=pl.get("created_at", ""),
            status=pl.get("status", "unknown"),
        ),
    }


@router.delete("/{document_id}")
async def delete_document(
    document_id: str,
    collection: str = KB_COLLECTION,
    current_user: User = Depends(get_current_user),
):
    # Delete all chunks
    chunk_ids = []
    points, _ = await vector_scroll(
        collection=collection,
        filters={"document_id": document_id, "type": "chunk"},
        limit=1000,
    )
    for p in points:
        chunk_ids.append(p["id"])

    for cid in chunk_ids:
        await delete_point(collection, cid)

    # Delete document metadata
    await delete_point(DOCUMENTS_INDEX, document_id)

    return {"success": True, "deleted_chunks": len(chunk_ids)}


@router.get("/{document_id}/chunks")
async def get_document_chunks(
    document_id: str,
    collection: str = KB_COLLECTION,
    limit: int = 100,
    offset: str | None = None,
    current_user: User = Depends(get_current_user),
):
    points, next_offset = await vector_scroll(
        collection=collection,
        filters={"document_id": document_id, "type": "chunk"},
        limit=limit,
        offset=offset,
    )
    chunks = []
    for p in points:
        pl = p["payload"]
        chunks.append(
            ChunkResponse(
                id=p["id"],
                document_id=pl.get("document_id", document_id),
                content=pl.get("content", ""),
                chunk_index=pl.get("chunk_index", 0),
            )
        )
    return {"success": True, "chunks": chunks, "next_offset": next_offset, "total": len(chunks)}


@router.post("/search")
async def search_documents(
    body: SearchRequest,
    current_user: User = Depends(get_current_user),
):
    results = await vector_search(
        collection=body.collection,
        query_text=body.query,
        limit=body.limit,
        score_threshold=body.score_threshold,
        filters=body.filters,
    )
    items = []
    for r in results:
        pl = r["payload"]
        chunk = ChunkResponse(
            id=r["id"],
            document_id=pl.get("document_id", ""),
            content=pl.get("content", ""),
            chunk_index=pl.get("chunk_index", 0),
            score=r["score"],
        )
        items.append({"chunk": chunk, "document": None})
    return {"success": True, "results": items, "total": len(items)}
